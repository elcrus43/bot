"""
Avito Price Parser - Квартиры
Парсер для сбора цен на квартиры с Avito (Selenium)
"""

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
from bs4 import BeautifulSoup
import pandas as pd
import json
import time
import random
import re
import logging
import argparse
from dataclasses import dataclass, field, asdict
from typing import List, Dict, Optional
from datetime import datetime
from tqdm import tqdm


# Коды регионов для Avito
REGIONS = {
    "Москва": "moskva",
    "Санкт-Петербург": "sankt-peterburg",
    "Новосибирск": "novosibirsk",
    "Екатеринбург": "ekaterinburg",
    "Казань": "kazan",
    "Нижний Новгород": "nizhnij_novgorod",
    "Челябинск": "chelyabinsk",
    "Самара": "samara",
    "Омск": "omsk",
    "Ростов-на-Дону": "rostov-na-donu",
    "Уфа": "ufa",
    "Красноярск": "krasnoyarsk",
    "Воронеж": "voronezh",
    "Пермь": "perm",
    "Волгоград": "volgograd",
    "Краснодар": "krasnodar",
    "Тула": "tula",
    "Тверь": "tver",
    "Сочи": "sochi",
    "Владивосток": "vladivostok",
    "Кировская область": "kirovskaya_oblast",
}


# Селекторы для парсинга (легко обновлять при изменении структуры Avito)
SELECTORS = {
    "item_card": {"data-marker": "item"},
    "item_card_alt": "article",
    "title": {"itemprop": "name"},
    "title_alt": ["h3", "a"],
    "price": {"itemprop": "price"},
    "price_meta": {"itemprop": "price"},
    "address_geo": "geo-text",
    "params_class": "item__params",
    "link_tag": "a",
}


@dataclass
class Flat:
    """Dataclass для хранения данных об объявлении"""
    title: str = ""
    price: int = 0
    price_raw: str = ""
    address: str = ""
    params: str = ""
    area: Optional[float] = None
    floor: Optional[str] = None
    url: str = ""
    region: str = ""
    parsed_at: str = ""
    is_valid: bool = True
    validation_errors: List[str] = field(default_factory=list)

    def to_dict(self) -> Dict:
        """Конвертация в словарь для сохранения"""
        d = asdict(self)
        d.pop("is_valid")
        d.pop("validation_errors")
        return d


class AvitoFlatParser:
    def __init__(self, region: str = "moskva", headless: bool = True):
        self.region = region
        self.headless = headless
        
        # Настройка логирования
        self.logger = logging.getLogger(__name__)
        if not self.logger.handlers:
            handler = logging.StreamHandler()
            formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
            handler.setFormatter(formatter)
            self.logger.addHandler(handler)
            self.logger.setLevel(logging.INFO)
        
        # Настройка WebDriver
        self.driver = self._init_driver()
        self.wait = WebDriverWait(self.driver, 15)

    def _init_driver(self) -> webdriver.Chrome:
        """Инициализация Chrome WebDriver"""
        options = Options()
        
        if self.headless:
            options.add_argument("--headless=new")
        
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option("useAutomationExtension", False)
        options.add_argument("--window-size=1920,1080")
        options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
        
        try:
            driver = webdriver.Chrome(options=options)
            # Скрытие webdriver-атрибутов
            driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
                "source": """
                    Object.defineProperty(navigator, 'webdriver', {get: () => undefined})
                """
            })
            self.logger.info("WebDriver инициализирован")
            return driver
        except WebDriverException as e:
            self.logger.error(f"Ошибка инициализации WebDriver: {e}")
            self.logger.info("Убедитесь, что ChromeDriver установлен: pip install webdriver-manager")
            raise

    def search_flats(self, rooms: str = "all", max_pages: int = 3) -> List[Flat]:
        """
        Поиск квартир по параметрам

        Args:
            rooms: Количество комнат ("1", "2", "3", "all")
            max_pages: Максимальное количество страниц

        Returns:
            Список объектов Flat с информацией о квартирах
        """
        all_flats: List[Flat] = []

        # Формируем URL
        if rooms == "all":
            url_base = f"https://www.avito.ru/{self.region}/kvartiry"
        else:
            url_base = f"https://www.avito.ru/{self.region}/kvartiry/komnat-{rooms}"

        for page in tqdm(range(1, max_pages + 1), desc="Страницы", unit="стр"):
            url = f"{url_base}?p={page}"

            try:
                self.logger.info(f"Загрузка страницы {page}/{max_pages}...")
                self.driver.get(url)
                
                # Ожидание загрузки контента
                try:
                    self.wait.until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, '[data-marker="item"], article'))
                    )
                except TimeoutException:
                    self.logger.warning("Таймаут ожидания элементов. Проверка на капчу...")
                    if self._is_captcha_page():
                        self.logger.error("Обнаружена капча. Подождите или решите вручную.")
                        # Даем время на ручное решение капчи
                        self.logger.info("Ожидание 30 секунд...")
                        time.sleep(30)
                        if self._is_captcha_page():
                            self.logger.error("Капча не решена. Остановка.")
                            break

                # Дополнительная прокрутка для загрузки контента
                self._scroll_page()

                # Получаем HTML после загрузки JS
                html = self.driver.page_source
                flats = self._parse_page(html)
                
                if not flats:
                    self.logger.info("Больше нет результатов на странице")
                    break

                all_flats.extend(flats)
                self.logger.info(f"Найдено {len(flats)} объявлений на странице {page}")

                # Адаптивная задержка
                time.sleep(random.uniform(3, 7))

            except WebDriverException as e:
                self.logger.error(f"Ошибка WebDriver: {e}")
                break

        self.logger.info(f"Парсинг завершён. Всего собрано {len(all_flats)} объявлений")
        return all_flats

    def _is_captcha_page(self) -> bool:
        """Проверка на наличие капчи"""
        try:
            page_source = self.driver.page_source.lower()
            return "captcha" in page_source or "подтвердите" in page_source or "robot" in page_source
        except:
            return False

    def _scroll_page(self):
        """Прокрутка страницы для загрузки контента"""
        try:
            scroll_pause = 0.5
            last_height = self.driver.execute_script("return document.body.scrollHeight")
            
            for _ in range(3):  # Максимум 3 прокрутки
                self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(scroll_pause)
                new_height = self.driver.execute_script("return document.body.scrollHeight")
                if new_height == last_height:
                    break
                last_height = new_height
            
            # Возврат наверх
            self.driver.execute_script("window.scrollTo(0, 0);")
            time.sleep(0.5)
        except Exception as e:
            self.logger.debug(f"Ошибка прокрутки: {e}")

    def close(self):
        """Закрытие WebDriver"""
        if self.driver:
            self.driver.quit()
            self.logger.info("WebDriver закрыт")

    def parse_single_listing(self, url: str) -> Optional[Flat]:
        """Парсинг конкретного объявления по URL"""
        self.logger.info(f"Загрузка объявления: {url}")
        self.driver.get(url)
        
        try:
            self.wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
            time.sleep(3)  # Ждем загрузки JS
            self._scroll_page()
            
            html = self.driver.page_source
            soup = BeautifulSoup(html, "html.parser")
            
            return self._extract_flat_from_page(soup, url)
        except Exception as e:
            self.logger.error(f"Ошибка парсинга объявления: {e}")
            return None

    def _extract_flat_from_page(self, soup, url: str) -> Optional[Flat]:
        """Извлечение данных объявления со страницы"""
        try:
            # Заголовок
            title = ""
            title_tag = soup.find("h1")
            if title_tag:
                title = title_tag.get_text(strip=True)
            
            # Цена
            price = 0
            price_raw = ""
            price_tag = soup.find("span", {"itemprop": "price"})
            if not price_tag:
                price_tag = soup.find("meta", {"itemprop": "price"})
            if price_tag:
                price_raw = price_tag.get("content", "").strip()
                if not price_raw:
                    price_raw = price_tag.get_text(strip=True)
                price_clean = price_raw.replace(" ", "").replace("₽", "").replace("руб", "")
                if price_clean.isdigit():
                    price = int(price_clean)
            
            # Адрес
            address = ""
            address_tag = soup.find("a", class_=re.compile(r"geo-link"))
            if address_tag:
                address = address_tag.get_text(strip=True)
            
            # Извлекаем параметры из заголовка (самый надежный способ)
            # Формат: "1-к. квартира, 30,8 м², 1/4 эт."
            area = None
            floor = None
            rooms = None
            
            # Количество комнат из заголовка
            rooms_match = re.search(r'(\d+)[-\s]комн', title, re.IGNORECASE)
            if rooms_match:
                rooms = int(rooms_match.group(1))
            
            # Площадь из заголовка (формат: 30,8 м²)
            area_match = re.search(r'(\d+[,.]?\d*)\s*м²', title)
            if area_match:
                area_str = area_match.group(1).replace(",", ".")
                area = float(area_str)
            
            # Этаж из заголовка (формат: 1/4 эт.)
            floor_match = re.search(r'(\d+)/\d+\s*эт', title)
            if floor_match:
                floor = floor_match.group(1)
            
            flat = Flat(
                title=title,
                price=price,
                price_raw=price_raw,
                address=address,
                area=area,
                floor=floor,
                url=url,
                region=self.region,
                parsed_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )
            
            # Сохраняем rooms в params для поиска похожих
            flat.params = f"{rooms}-комн, {area} м², {floor} этаж" if rooms else f"{area} м², {floor} этаж"
            
            self._validate_flat(flat)
            return flat
            
        except Exception as e:
            self.logger.error(f"Ошибка извлечения данных: {e}")
            return None

    def find_similar_listings(self, target_flat: Flat, max_pages: int = 3) -> List[Flat]:
        """Поиск похожих объявлений"""
        self.logger.info("Поиск похожих объявлений...")
        
        # Извлекаем параметры для поиска
        rooms_match = re.search(r'(\d+)-комн', target_flat.params)
        rooms = rooms_match.group(1) if rooms_match else "all"
        
        # Формируем URL поиска
        if rooms == "all":
            url_base = f"https://www.avito.ru/{self.region}/kvartiry"
        else:
            url_base = f"https://www.avito.ru/{self.region}/kvartiry/komnat-{rooms}"
        
        similar_flats: List[Flat] = []
        
        for page in tqdm(range(1, max_pages + 1), desc="Поиск похожих", unit="стр"):
            url = f"{url_base}?p={page}"
            
            try:
                self.driver.get(url)
                
                try:
                    self.wait.until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, '[data-marker="item"], article'))
                    )
                except TimeoutException:
                    self.logger.warning("Таймаут ожидания элементов")
                    break
                
                self._scroll_page()
                html = self.driver.page_source
                flats = self._parse_page(html)
                
                if not flats:
                    break
                
                # Фильтруем похожие объявления (исключаем аренду)
                for flat in flats:
                    if self._is_similar(flat, target_flat) and not self._is_rental(flat):
                        similar_flats.append(flat)

                similar_count = len([f for f in flats if self._is_similar(f, target_flat) and not self._is_rental(f)])
                self.logger.info(f"На странице {page}: {similar_count} похожих (без аренды)")
                time.sleep(random.uniform(3, 7))
                
            except WebDriverException as e:
                self.logger.error(f"Ошибка WebDriver: {e}")
                break
        
        self.logger.info(f"Найдено {len(similar_flats)} похожих объявлений")
        return similar_flats

    def _is_similar(self, flat: Flat, target: Flat) -> bool:
        """Проверка похожести объявлений"""
        # Сравниваем площадь (±50%)
        if flat.area and target.area:
            area_diff = abs(flat.area - target.area) / target.area
            if area_diff > 0.5:
                return False
        
        # Проверяем этаж (±5 этажей)
        if flat.floor and target.floor:
            try:
                floor_diff = abs(int(flat.floor) - int(target.floor))
                if floor_diff > 5:
                    return False
            except:
                pass
        
        return True

    def compare_prices(self, target_flat: Flat, similar_flats: List[Flat]):
        """Сравнительный анализ цен"""
        if not similar_flats:
            print("Нет похожих объявлений для сравнения")
            return
        
        # Фильтруем выбросы (цена > 100,000 и < 100 млн) — убираем аренду
        valid_flats = [f for f in similar_flats if 100_000 < f.price < 100_000_000]
        prices = [f.price for f in valid_flats]
        
        if not prices:
            print("Не удалось извлечь цены для сравнения")
            return
        
        target_price = target_flat.price
        avg_price = sum(prices) / len(prices)
        sorted_prices = sorted(prices)
        median_price = sorted_prices[len(sorted_prices) // 2]
        
        # Фильтруем по цене (убираем экстремумы, 10-90 перцентиль)
        p10_idx = max(0, int(len(sorted_prices) * 0.10))
        p90_idx = min(len(sorted_prices) - 1, int(len(sorted_prices) * 0.90))
        filtered_prices = sorted_prices[p10_idx:p90_idx + 1]
        
        if filtered_prices:
            avg_filtered = sum(filtered_prices) / len(filtered_prices)
            median_filtered = filtered_prices[len(filtered_prices) // 2]
        else:
            avg_filtered = avg_price
            median_filtered = median_price
        
        # Цена за м²
        target_price_m2 = target_price / target_flat.area if target_flat.area else 0
        prices_m2 = [f.price / f.area for f in valid_flats if f.area and f.area > 0]
        avg_price_m2 = sum(prices_m2) / len(prices_m2) if prices_m2 else 0
        
        # Фильтруем цену за м² (10-90 перцентиль)
        sorted_prices_m2 = sorted(prices_m2) if prices_m2 else []
        if sorted_prices_m2:
            p10_m2 = max(0, int(len(sorted_prices_m2) * 0.10))
            p90_m2 = min(len(sorted_prices_m2) - 1, int(len(sorted_prices_m2) * 0.90))
            filtered_prices_m2 = sorted_prices_m2[p10_m2:p90_m2 + 1]
            avg_price_m2_filtered = sum(filtered_prices_m2) / len(filtered_prices_m2) if filtered_prices_m2 else avg_price_m2
        else:
            avg_price_m2_filtered = avg_price_m2
        
        print("\n" + "="*60)
        print("📊 СРАВНИТЕЛЬНЫЙ АНАЛИЗ ЦЕН")
        print("="*60)
        
        print(f"\n🏠 Ваше объявление:")
        print(f"   Заголовок: {target_flat.title}")
        print(f"   Цена: {target_price:,} ₽")
        if target_flat.area:
            print(f"   Площадь: {target_flat.area} м²")
            print(f"   Цена за м²: {target_price_m2:,.0f} ₽/м²")
        
        print(f"\n📈 Рынок ({len(valid_flats)} объявлений, без аренды и выбросов):")
        print(f"   Средняя цена: {avg_filtered:,.0f} ₽")
        print(f"   Медиана: {median_filtered:,} ₽")
        print(f"   Мин: {min(filtered_prices):,} ₽")
        print(f"   Макс: {max(filtered_prices):,} ₽")
        
        if avg_price_m2_filtered > 0:
            print(f"   Средняя цена за м²: {avg_price_m2_filtered:,.0f} ₽/м²")
        
        print(f"\n💡 Вывод:")
        # Используем медиану для сравнения (более устойчива к выбросам)
        if target_price > median_filtered * 1.15:
            diff = (target_price - median_filtered) / median_filtered * 100
            print(f"   ⚠️ Цена ВЫШЕ рынка на {diff:.1f}%")
            print(f"   Рекомендуемая цена: {median_filtered:,.0f} ₽")
        elif target_price < median_filtered * 0.85:
            diff = (median_filtered - target_price) / median_filtered * 100
            print(f"   ✅ Цена НИЖЕ рынка на {diff:.1f}%")
            print(f"   Хорошее предложение!")
        else:
            print(f"   ✓ Цена соответствует рынку")
        
        # Рекомендуемая цена
        if avg_price_m2_filtered > 0 and target_flat.area:
            recommended_price = avg_price_m2_filtered * target_flat.area
            print(f"\n💰 Рекомендуемая цена: {recommended_price:,.0f} ₽")
            print(f"   (на основе средней цены за м² × площадь)")
        
        print("="*60)

    def _parse_page(self, html: str) -> List[Flat]:
        """Парсинг HTML страницы"""
        soup = BeautifulSoup(html, "html.parser")
        flats: List[Flat] = []

        # Находим все карточки объявлений
        cards = soup.find_all("div", SELECTORS["item_card"])

        if not cards:
            # Альтернативный селектор
            cards = soup.find_all(SELECTORS["item_card_alt"])

        for card in cards:
            flat = self._extract_flat_data(card)
            if flat:
                flats.append(flat)

        return flats

    def _extract_flat_data(self, card) -> Optional[Flat]:
        """Извлечение данных из карточки квартиры"""
        try:
            # Заголовок — ищем ссылку с заголовком объявления
            title = ""
            title_tag = card.find("span", SELECTORS["title"])
            
            # Пробуем разные селекторы
            if not title_tag:
                title_tag = card.find("a", class_=re.compile(r"title|name|snippet"))
            
            if not title_tag:
                # Ищем любую ссылку с текстом, похожим на заголовок
                for a_tag in card.find_all("a", href=True):
                    text = a_tag.get_text(strip=True)
                    if text and len(text) > 10 and "фото" not in text.lower():
                        title_tag = a_tag
                        break
            
            if title_tag:
                title = title_tag.get_text(strip=True)
            
            # Пропускаем если заголовок не похож на объявление
            if not title or len(title) < 5 or "фото" in title.lower():
                return None
            
            # Цена
            price_raw = ""
            price = 0
            price_tag = card.find("span", SELECTORS["price"])
            if not price_tag:
                price_tag = card.find("meta", SELECTORS["price_meta"])
            
            if not price_tag:
                # Ищем цену по тексту
                for span in card.find_all("span"):
                    text = span.get_text(strip=True)
                    if "₽" in text or "руб" in text:
                        price_tag = span
                        break

            if price_tag:
                price_raw = price_tag.get("content", "").strip()
                if not price_raw:
                    price_raw = price_tag.get_text(strip=True)

                # Парсинг цены
                price_clean = price_raw.replace(" ", "").replace("₽", "").replace("руб", "").replace("₽", "")
                if price_clean.isdigit():
                    price = int(price_clean)

            # Адрес
            address = ""
            address_tag = card.find("div", string=lambda text: text and "ул." in text)
            if not address_tag:
                address_tag = card.find("span", class_=SELECTORS["address_geo"])
            if address_tag:
                address = address_tag.get_text(strip=True)

            # Параметры (этаж, площадь)
            params = ""
            params_tag = card.find("div", class_=SELECTORS["params_class"])
            if not params_tag:
                # Ищем параметры по тексту с м²
                for div in card.find_all("div"):
                    text = div.get_text(strip=True)
                    if "м²" in text:
                        params_tag = div
                        break
            if params_tag:
                params = params_tag.get_text(strip=True)

            # Извлечение площади из параметров
            area = None
            area_match = re.search(r'(\d+\.?\d*)\s*м²', params)
            if area_match:
                area = float(area_match.group(1))
            
            # Если не нашли в params, ищем в заголовке
            if not area:
                area_match = re.search(r'(\d+[,.]?\d*)\s*м²', title)
                if area_match:
                    area = float(area_match.group(1).replace(",", "."))

            # Извлечение этажа
            floor = None
            floor_match = re.search(r'(\d+)\s*этаж', params)
            if floor_match:
                floor = floor_match.group(1)
            
            # Если не нашли в params, ищем в заголовке
            if not floor:
                floor_match = re.search(r'(\d+)/\d+\s*эт', title)
                if floor_match:
                    floor = floor_match.group(1)

            # Ссылка
            link_tag = card.find(SELECTORS["link_tag"], href=True)
            link = link_tag["href"] if link_tag else ""
            if link and not link.startswith("http"):
                link = f"https://www.avito.ru{link}"

            if not title:
                return None

            flat = Flat(
                title=title,
                price=price,
                price_raw=price_raw,
                address=address,
                params=params,
                area=area,
                floor=floor,
                url=link,
                region=self.region,
                parsed_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )

            # Валидация
            self._validate_flat(flat)
            return flat

        except Exception as e:
            self.logger.error(f"Ошибка при парсинге карточки: {e}")
            return None

    def _is_rental(self, flat: Flat) -> bool:
        """Проверка, является ли объявление арендой"""
        rent_keywords = ["аренда", "сдам", "снять", "посуточно", "долгосрок", "на месяц", "месяц"]
        text = (flat.title + " " + flat.params).lower()
        return any(kw in text for kw in rent_keywords)

    def _validate_flat(self, flat: Flat) -> None:
        """Валидация данных объявления"""
        errors = []

        if not flat.title:
            errors.append("Отсутствует название")

        if flat.price <= 0:
            errors.append("Цена не указана или некорректна")

        if flat.price > 1_000_000_000:
            errors.append("Цена превышает 1 млрд — возможно ошибка парсинга")

        if flat.area and (flat.area <= 0 or flat.area > 1000):
            errors.append("Площадь некорректна")

        if not flat.url:
            errors.append("Отсутствует ссылка")

        flat.validation_errors = errors
        flat.is_valid = len(errors) == 0

    def save_to_csv(self, flats: List[Flat], filename: str = "avito_flats.csv"):
        """Сохранение данных в CSV"""
        if not flats:
            self.logger.warning("Нет данных для сохранения")
            return

        data = [f.to_dict() for f in flats]
        df = pd.DataFrame(data)
        df.to_csv(filename, index=False, encoding="utf-8-sig")
        self.logger.info(f"Данные сохранены в {filename}")

    def save_to_json(self, flats: List[Flat], filename: str = "avito_flats.json"):
        """Сохранение данных в JSON"""
        if not flats:
            self.logger.warning("Нет данных для сохранения")
            return

        data = [f.to_dict() for f in flats]
        with open(filename, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        self.logger.info(f"Данные сохранены в {filename}")

    def save_to_excel(self, flats: List[Flat], filename: str = "avito_flats.xlsx"):
        """Сохранение данных в Excel с форматированием"""
        if not flats:
            self.logger.warning("Нет данных для сохранения")
            return

        data = [f.to_dict() for f in flats]
        df = pd.DataFrame(data)

        # Создаём Excel с форматированием
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Квартиры")
            
            ws = writer.sheets["Квартиры"]
            
            # Форматирование заголовков
            from openpyxl.styles import Font, PatternFill, Alignment
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            
            # Автоширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        self.logger.info(f"Данные сохранены в {filename}")

    def print_analytics(self, flats: List[Flat]):
        """Вывод базовой аналитики по квартирам"""
        if not flats:
            self.logger.info("Нет данных для аналитики")
            return

        # Фильтруем только валидные объявления с ценой
        valid_flats = [f for f in flats if f.is_valid and f.price > 0]
        prices = [f.price for f in valid_flats]

        if prices:
            print("\n=== Аналитика цен на квартиры ===")
            print(f"Всего объявлений: {len(flats)}")
            print(f"Валидных с ценой: {len(prices)}")
            print(f"Мин. цена: {min(prices):,} ₽")
            print(f"Макс. цена: {max(prices):,} ₽")
            print(f"Средняя цена: {sum(prices)//len(prices):,} ₽")
            print(f"Медиана: {sorted(prices)[len(prices)//2]:,} ₽")

            # Цена за м²
            prices_per_m2 = []
            for flat in valid_flats:
                if flat.area and flat.area > 0:
                    price_per_m2 = flat.price / flat.area
                    prices_per_m2.append(price_per_m2)

            if prices_per_m2:
                print(f"\nЦена за м²:")
                print(f"  Мин: {int(min(prices_per_m2)):,} ₽/м²")
                print(f"  Макс: {int(max(prices_per_m2)):,} ₽/м²")
                print(f"  Средняя: {int(sum(prices_per_m2)//len(prices_per_m2)):,} ₽/м²")
        else:
            self.logger.info("Не удалось извлечь цены для аналитики")


def main():
    """Основная функция"""
    parser_arg = argparse.ArgumentParser(description="Парсер цен на квартиры Avito")
    parser_arg.add_argument("-r", "--region", type=str, default=None, help="Регион (название или код)")
    parser_arg.add_argument("-c", "--rooms", type=str, default=None, choices=["1", "2", "3", "all"], help="Количество комнат")
    parser_arg.add_argument("-p", "--pages", type=int, default=None, help="Количество страниц")
    parser_arg.add_argument("-o", "--output", type=str, default="avito_flats", help="Префикс имени файла")
    parser_arg.add_argument("-v", "--verbose", action="store_true", help="Подробное логирование")
    parser_arg.add_argument("--visible", action="store_true", help="Показать браузер (не headless)")
    parser_arg.add_argument("--url", type=str, default=None, help="URL объявления Avito для анализа")
    parser_arg.add_argument("--similar", type=int, default=3, help="Количество страниц для поиска похожих")
    
    args = parser_arg.parse_args()

    # Настройка логирования
    log_level = logging.DEBUG if args.verbose else logging.INFO
    logging.basicConfig(
        level=log_level,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    
    print("=== Парсер цен на квартиры Avito ===\n")

    # Режим анализа конкретного объявления
    if args.url:
        parser = AvitoFlatParser(headless=not args.visible)
        try:
            # Извлекаем регион из URL
            url_region = args.url.split("/avito.ru/")[1].split("/")[0] if "/avito.ru/" in args.url else None
            if url_region:
                parser.region = url_region
            
            # Парсим объявление
            print("📋 Парсинг объявления...")
            target_flat = parser.parse_single_listing(args.url)
            
            if not target_flat:
                print("❌ Не удалось распарсить объявление")
                return
            
            print(f"\n✅ Объявление найдено:")
            print(f"   Заголовок: {target_flat.title}")
            print(f"   Цена: {target_flat.price:,} ₽")
            print(f"   Площадь: {target_flat.area} м²" if target_flat.area else "   Площадь: не указана")
            print(f"   Этаж: {target_flat.floor}" if target_flat.floor else "   Этаж: не указан")
            print(f"   Адрес: {target_flat.address}")
            
            # Сохраняем целевое объявление
            parser.save_to_json([target_flat], f"{args.output}_target.json")
            print(f"\n💾 Данные сохранены: {args.output}_target.json")
            
            # Поиск похожих
            print(f"\n🔍 Поиск похожих объявлений ({args.similar} страниц)...")
            similar_flats = parser.find_similar_listings(target_flat, max_pages=args.similar)
            
            if similar_flats:
                # Сохраняем похожие
                parser.save_to_csv(similar_flats, f"{args.output}_similar.csv")
                parser.save_to_excel(similar_flats, f"{args.output}_similar.xlsx")
                print(f"💾 Найдено {len(similar_flats)} похожих объявлений")
                
                # Сравнительный анализ
                parser.compare_prices(target_flat, similar_flats)
            else:
                print("⚠️ Похожих объявлений не найдено")
                
        finally:
            parser.close()
        return

    # Определение региона
    region_code = args.region
    if not region_code:
        print("Доступные регионы:")
        for i, (name, code) in enumerate(REGIONS.items(), 1):
            print(f"  {i}. {name}")
        region_choice = input("\nВыберите регион (номер или название): ").strip()
        if region_choice.isdigit():
            idx = int(region_choice) - 1
            if 0 <= idx < len(REGIONS):
                region_code = list(REGIONS.values())[idx]
        else:
            for name, code in REGIONS.items():
                if region_choice.lower() in name.lower():
                    region_code = code
                    break
            else:
                region_code = "moskva"

    # Определение комнат
    rooms = args.rooms
    if not rooms:
        rooms = input("\nКоличество комнат (1/2/3/all): ").strip().lower()
        if rooms not in ["1", "2", "3"]:
            rooms = "all"

    # Определение страниц
    max_pages = args.pages
    if max_pages is None:
        max_pages_input = input("Количество страниц (по умолчанию 3): ").strip()
        max_pages = int(max_pages_input) if max_pages_input.isdigit() else 3

    # Парсинг
    parser = AvitoFlatParser(region_code, headless=not args.visible)
    print(f"\nНачинаем поиск квартир в регионе: {region_code}")
    print(f"Комнат: {rooms}, Страниц: {max_pages}\n")
    
    try:
        flats = parser.search_flats(rooms, max_pages)

        if flats:
            # Сохраняем
            csv_file = f"{args.output}.csv"
            json_file = f"{args.output}.json"
            xlsx_file = f"{args.output}.xlsx"
            
            parser.save_to_csv(flats, csv_file)
            parser.save_to_json(flats, json_file)
            parser.save_to_excel(flats, xlsx_file)

            # Аналитика
            parser.print_analytics(flats)

            print(f"\nГотово! Спарсено {len(flats)} объявлений")
            print(f"Файлы: {csv_file}, {json_file}, {xlsx_file}")
        else:
            print("Ничего не найдено")
    finally:
        parser.close()


if __name__ == "__main__":
    main()
